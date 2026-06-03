"""
Convert TT23 Procedure Codes Excel to JSON database.
Output: D:\MediVoice_VN\data\reference\tt23_procedures.json
"""

import openpyxl
import json
import os
import re

DATA_DIR = r'D:\MediVoice_VN\data\reference'
src_file = os.path.join(DATA_DIR, 'TT23_procedure_codes.xlsx')

if not os.path.exists(src_file):
    print(f"ERROR: File not found: {src_file}")
    exit(1)

print(f"Source: {src_file}")

# --- Parse Excel to JSON ---
wb = openpyxl.load_workbook(src_file)
print(f"Sheets ({len(wb.sheetnames)}): {wb.sheetnames[:5]}...")


def extract_keywords(name: str, specialty: str) -> list:
    keywords = []
    name_lower = name.lower()
    keywords.append(name_lower)
    words = re.split(r'[\s\(\),\-/\[\]]+', name_lower)
    keywords.extend([w for w in words if len(w) >= 4])
    if 'siêu âm' in name_lower:
        keywords.extend(['sa', 'ultrasound', 'sieu am'])
    if 'x-quang' in name_lower or 'chụp x' in name_lower:
        keywords.extend(['xquang', 'xray', 'x quang'])
    if 'cắt lớp vi tính' in name_lower:
        keywords.extend(['ct', 'clvt', 'ct scan'])
    if 'cộng hưởng từ' in name_lower:
        keywords.extend(['mri', 'cong huong tu'])
    if 'doppler' in name_lower:
        keywords.append('doppler')
    return list(set(kw for kw in keywords if kw.strip()))


db = {
    "_meta": {
        "source": "Thong tu 23 - Phu luc 02",
        "title": "Danh muc ky thuat thuc hien tu ngay 01/7/2026",
        "total_specialties": len(wb.sheetnames),
        "sheet_names": wb.sheetnames
    },
    "by_code": {},
    "by_specialty": {},
    "keyword_index": {}
}

total = 0

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    specialty_procedures = []

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        if not row or not any(row):
            continue
        code = row[0]
        name = row[4] if len(row) > 4 else None
        specialty = row[2] if len(row) > 2 else sheet_name
        protocol_ref = row[3] if len(row) > 3 else None

        if not isinstance(code, (int, float)):
            continue
        if not name or not isinstance(name, str):
            continue

        code_str = str(int(code))
        name_str = str(name).strip()
        specialty_str = str(specialty).strip() if specialty else sheet_name
        protocol_str = str(protocol_ref).strip() if protocol_ref else ""
        keywords = extract_keywords(name_str, specialty_str)

        procedure = {
            "code": code_str,
            "name": name_str,
            "specialty": specialty_str,
            "protocol_ref": protocol_str,
            "keywords": keywords
        }

        db["by_code"][code_str] = procedure
        specialty_procedures.append(code_str)

        for kw in keywords:
            if kw not in db["keyword_index"]:
                db["keyword_index"][kw] = []
            if code_str not in db["keyword_index"][kw]:
                db["keyword_index"][kw].append(code_str)

        total += 1

    if specialty_procedures:
        db["by_specialty"][sheet_name] = specialty_procedures

db["_meta"]["total_procedures"] = total

# Save full JSON
out_json = os.path.join(DATA_DIR, 'tt23_procedures.json')
with open(out_json, 'w', encoding='utf-8') as f:
    json.dump(db, f, ensure_ascii=False, indent=2)
print(f"\nFull JSON: {out_json} ({os.path.getsize(out_json)/1024:.0f} KB)")
print(f"Procedures: {total} | Specialties: {len(db['by_specialty'])} | Keywords: {len(db['keyword_index'])}")

# Save CDHA-only JSON
cdha_sheet = '23.ĐIỆN QUANG'
if cdha_sheet in db["by_specialty"]:
    cdha_codes = db["by_specialty"][cdha_sheet]
    cdha_db = {
        "_meta": {"source": "TT23", "specialty": cdha_sheet, "count": len(cdha_codes)},
        "procedures": {code: db["by_code"][code] for code in cdha_codes},
        "by_type": {"sieu_am": [], "x_quang": [], "ct_scan": [], "mri": [], "other": []}
    }
    for code in cdha_codes:
        name = db["by_code"][code]["name"].lower()
        if 'siêu âm' in name:
            cdha_db["by_type"]["sieu_am"].append(code)
        elif 'x-quang' in name or 'chụp phim' in name or 'cephalometric' in name:
            cdha_db["by_type"]["x_quang"].append(code)
        elif 'cắt lớp vi tính' in name:
            cdha_db["by_type"]["ct_scan"].append(code)
        elif 'cộng hưởng từ' in name:
            cdha_db["by_type"]["mri"].append(code)
        else:
            cdha_db["by_type"]["other"].append(code)

    cdha_json = os.path.join(DATA_DIR, 'tt23_cdha.json')
    with open(cdha_json, 'w', encoding='utf-8') as f:
        json.dump(cdha_db, f, ensure_ascii=False, indent=2)
    print(f"\nCDHA JSON: {cdha_json} ({os.path.getsize(cdha_json)/1024:.0f} KB)")
    print(f"  Siêu âm: {len(cdha_db['by_type']['sieu_am'])}")
    print(f"  X-quang: {len(cdha_db['by_type']['x_quang'])}")
    print(f"  CT scan: {len(cdha_db['by_type']['ct_scan'])}")
    print(f"  MRI:     {len(cdha_db['by_type']['mri'])}")
    print(f"  Other:   {len(cdha_db['by_type']['other'])}")

# Sample lookups
print("\n--- SAMPLE RECORDS ---")
samples = ["6408", "6462", "6619"]
for code in samples:
    if code in db["by_code"]:
        p = db["by_code"][code]
        print(f"  [{p['code']}] {p['name'][:70]}")

print("\nDone.")
