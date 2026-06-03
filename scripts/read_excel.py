import openpyxl
import glob
import os
import sys

# Find the file
docs_dir = r'D:\MediVoice_VN\docs'
files = [f for f in os.listdir(docs_dir) if 'PHU LUC' in f or 'CDHA' in f.upper() or 'CĐHA' in f]
if not files:
    print("File not found")
    sys.exit(1)

filepath = os.path.join(docs_dir, files[0])
print(f"Reading: {files[0]}")

wb = openpyxl.load_workbook(filepath)
print(f"Sheets: {wb.sheetnames}\n")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"=== SHEET: {sheet_name} ({ws.max_row} rows x {ws.max_column} cols) ===")
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 150), values_only=True)):
        if any(cell is not None for cell in row):
            # Clean up None values for display
            clean = [str(c) if c is not None else '' for c in row]
            print(f"  R{i+1}: {' | '.join(c for c in clean if c)}")
    print()
