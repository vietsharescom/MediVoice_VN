import openpyxl
import os

docs_dir = r'D:\MediVoice_VN\docs'
files = [f for f in os.listdir(docs_dir) if 'PHU LUC' in f]
filepath = os.path.join(docs_dir, files[0])

wb = openpyxl.load_workbook(filepath)
print(f"File: {files[0]}")
print(f"Tong so sheet: {len(wb.sheetnames)}")
print(f"Tat ca sheet: {wb.sheetnames}\n")

# Read DIEN QUANG sheet (radiology/imaging = CDHA)
target_sheets = ['23.ĐIỆN QUANG', '22.KT CHUNG']
for target in target_sheets:
    if target in wb.sheetnames:
        ws = wb[target]
        print(f"\n{'='*70}")
        print(f"SHEET: {target} ({ws.max_row} rows)")
        print('='*70)
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True)):
            if any(cell is not None for cell in row):
                clean = [str(c).strip() if c is not None else '' for c in row]
                non_empty = [c for c in clean if c]
                if non_empty:
                    print(f"  {' | '.join(non_empty)}")

# Also show first 5 rows of all sheets to understand structure
print("\n\n" + "="*70)
print("TOM TAT: 5 DONG DAU TUNG SHEET")
print("="*70)
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n--- {sheet_name} ---")
    count = 0
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        if any(cell is not None for cell in row):
            clean = [str(c).strip() if c is not None else '' for c in row]
            non_empty = [c for c in clean if c]
            if non_empty:
                print(f"  {' | '.join(non_empty[:4])}")
                count += 1
                if count >= 4:
                    print(f"  ... ({ws.max_row} rows total)")
                    break
